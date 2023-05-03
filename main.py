import openpyxl
from datetime import datetime, timedelta
import calendar
import holidays
import os
import configparser
import wget
from rich.console import Console
from rich.theme import Theme
from rich.table import Table


# Run this to COMPILE:
# pyinstaller --name=EvidencijaRada --noconfirm --onefile --console --icon=C:\Users\msemes\Documents\Python\PriceParser\icon.ico main.py


# Console Themeing
custom_theme = Theme({"success": "bold bright_green", "error": "bold red", "warning": "bold dark_orange3", "accent": "bold cyan", "general": "bold grey35", "cyan": "bold cyan"})
console = Console(theme=custom_theme)


# Dates and Times
now = datetime.now()  # Get the current date and time
todays_date = now.strftime("%d.%m.%Y")
last_month = now - timedelta(days=30)  # Subtract 30 days to get the previous month
last_month_year = last_month.year  # Get the year component of the previous month's date
current_month = int(datetime.now().month)
previous_month = int(datetime.now().month) - 1


# Document template name

DOCUMENT_TEMPLATE = 'evidencija_o_radnom_template.xlsx'
if not os.path.exists(DOCUMENT_TEMPLATE):
    console.print("Template does not exist.", style="warning")
else:
    console.print("Template exists, updating.", style="warning")
    os.remove(DOCUMENT_TEMPLATE)

URL = "https://github.com/mariosemes/Evidencija-Radar/raw/main/evidencija_o_radnom_template.xlsx"
console.print("Downloading template file.", style="warning")
response = wget.download(URL, DOCUMENT_TEMPLATE)
console.print("Template downloaded.", style="success")

if not os.path.exists(DOCUMENT_TEMPLATE):
    console.print("Something went wrong. Can't download xlsx file.", style="error")
    input("Press Enter to continue...")


# Config
# check if the config file exists, if not, create one
config_file = 'config.ini'
while not os.path.exists(config_file):
    console.print("Config file missing.", style="error")
    with open(config_file, 'w', encoding='utf-8') as f:
        f.write('[radno_vrijeme]\n')
        f.write('rv_pocetak = 9\n')
        f.write('rv_kraj = 17\n')
        f.write('[radnik]\n')
        f.write('zaposlenik = \n')
    console.print("Creating config file.", style="warning")
    if os.path.exists(config_file):
        console.print("Config file created.", style="success")
        console.print("[warning]--[/]")
        console.print("[warning]--  Prvi puta pokrećete aplikaciju![/]")
        console.print("[warning]--[/]  Ako ne želite uzastupno upisivati ime Zaposlenika,")
        console.print("[warning]--[/]  potrebno je pokrenuti novoizrađenu datoteku [cyan]config.ini[/]")
        console.print("[warning]--[/]  te pod stavku zaposlenik unesite ime i prezime koje zelite.")
        console.print("[warning]--[/]  npr. [cyan]zaposlenik = Ivan Horvat[/]")
        console.print("[warning]--[/]")
        input("Press Enter to continue...")
    else:
        console.print("Issue creating config file.", style="error")
        input("Press Enter to continue...")
        exit()

# read the custom variables from the config file
config = configparser.ConfigParser()
config.read(config_file, encoding='utf-8')
rv_kraj = int(config.get('radno_vrijeme', 'rv_kraj'))
rv_pocetak = int(config.get('radno_vrijeme', 'rv_pocetak'))
rv_sveukupno = int(rv_kraj) - int(rv_pocetak)
# Converts integer into time format like 8 = 08:00
# time_str = '{:02d}:00'.format(rv_poceta).zfill(5)
zaposlenik = config.get('radnik', 'zaposlenik')


def month_converter(month):
    MONTH_LIST = ['Siječanj', 'Veljača', 'Ožujak', 'Travanj', 'Svibanj', 'Lipanj', 'Srpanj', 'Kolovoz', 'Rujan',
                  'Listopad', 'Studeni', 'Prosinac']
    month = month - 1
    month_converted = MONTH_LIST[month]
    return month_converted


def clean_holidays(document, holidays_list):
    workbook = openpyxl.load_workbook(document)

    # Access the active worksheet
    worksheet = workbook.active

    # Saturday
    for h in holidays_list:
        h = h + 5
        colB = "B" + str(h)
        colC = "C" + str(h)
        colD = "D" + str(h)
        colH = "H" + str(h)

        # Write to the worksheet
        worksheet[colB] = ''
        worksheet[colC] = ''
        worksheet[colD] = ''
        worksheet[colH] = 'BLA'

    # Save the workbook to a file
    workbook.save(document)

def clean_weekends(document, sunday_dates):
    workbook = openpyxl.load_workbook(document)

    # Access the active worksheet
    worksheet = workbook.active

    # Saturday
    for s in sunday_dates:
        s = s + 4
        colB = "B" + str(s)
        colC = "C" + str(s)
        colD = "D" + str(s)
        colH = "H" + str(s)

        # Write to the worksheet
        worksheet[colB] = ''
        worksheet[colC] = ''
        worksheet[colD] = ''
        worksheet[colH] = 'TO'

    # Sunday
    for s in sunday_dates:
        s = s + 5
        colB = "B" + str(s)
        colC = "C" + str(s)
        colD = "D" + str(s)
        colH = "H" + str(s)

        # Write to the worksheet
        worksheet[colB] = ''
        worksheet[colC] = ''
        worksheet[colD] = ''
        worksheet[colH] = 'NED'

    # Save the workbook to a file
    workbook.save(document)



def get_holidays(year, month):
    # get the holidays for Croatia in the specified year
    hr_holidays = holidays.CountryHoliday('HR', years=[year])
    holidays_list = []

    # loop through the holidays and filter for the desired month
    for holiday_date, name in hr_holidays.items():
        if holiday_date.year == year and holiday_date.month == month:
            holidays_list.append(holiday_date.day)
            print(holiday_date.day, name)

    return holidays_list


def get_sundays(year, month):
    # Get the number of days in the given month
    num_days = calendar.monthrange(year, month)[1]

    # Initialize an empty list to store the Sunday dates
    sunday_dates = []

    # Loop over each day in the month
    for day in range(1, num_days + 1):
        # Get the weekday for the current day (0=Monday, 6=Sunday)
        weekday = calendar.weekday(year, month, day)

        # If the weekday is Sunday (6), add the date to the list of Sunday dates
        if weekday == 6:
            sunday_date = day
            sunday_dates.append(sunday_date)

    #print("Vikend za odabrani mjesec:")
    #for s in sunday_dates:
    #    print(str(s-1) + "." + str(month) + "." + str(year)+" - Subota")
    #    print(str(s)+"."+str(month)+"."+str(year)+" - Nedjelja")

    return sunday_dates


def count_ukupno(document, working_day_hours):
    workbook = openpyxl.load_workbook(document)

    # Access the active worksheet
    worksheet = workbook.active

    # Redovan Rad
    populated_fields = 0
    for row in worksheet.iter_rows(min_row=6, max_row=36, min_col=4, max_col=4):
        for cell in row:
            if cell.value is not None:
                populated_fields += 1

    if populated_fields > 0:
        populated_fields_multiplier = populated_fields * working_day_hours
        # Write to the worksheet
        worksheet['D37'] = populated_fields_multiplier
    else:
        worksheet['D37'] = '-'

    # Ostalo
    populated_fields = 0
    for row in worksheet.iter_rows(min_row=6, max_row=36, min_col=7, max_col=7):
        for cell in row:
            if cell.value is not None:
                populated_fields += 1

    if populated_fields > 0:
        populated_fields_multiplier = populated_fields * working_day_hours
        # Write to the worksheet
        worksheet['G37'] = populated_fields_multiplier
    else:
        worksheet['G37'] = '-'

    # Save the workbook to a file
    workbook.save(document)

def content_filler(document):
    workbook = openpyxl.load_workbook(document)

    # Access the active worksheet
    worksheet = workbook.active

    counter = 6
    while counter <= 36:
        colB = "B" + str(counter)
        colC = "C" + str(counter)
        colD = "D" + str(counter)

        # Write to the worksheet
        worksheet[colB] = '{:02d}:00'.format(rv_pocetak).zfill(5)
        worksheet[colC] = '{:02d}:00'.format(rv_kraj).zfill(5)
        worksheet[colD] = '{:02d}:00'.format(rv_sveukupno).zfill(5)
        counter += 1

    # Save the workbook to a file
    workbook.save(document)

    return rv_sveukupno


def insert_custom(document, choice):
    workbook = openpyxl.load_workbook(document)

    # Access the active worksheet
    worksheet = workbook.active

    # Clearing console
    os.system('cls' if os.name == 'nt' else 'clear')

    console.print("[warning]!!! Potrebno je unijeti samo dan kao broj bez mjeseca i godine (npr. dvadeset i treći = 23) !!![/]")
    while True:
        date_from = input("Početak: ")
        if not date_from.isdigit():
            console.print("[error]Niste unijeli broj. Pokušajte ponovo.[/]")
        elif int(date_from) > 32:
            console.print("[error]Broj je visok. Pokušajte ponovo.[/]")
        else:
            break
    while True:
        date_to = input("Kraj: ")
        if not date_to.isdigit():
            console.print("[error]Niste unijeli broj. Pokušajte ponovo.[/]")
        elif int(date_to) > 32:
            console.print("[error]Broj je visok. Pokušajte ponovo.[/]")
        elif int(date_to) < int(date_from):
            console.print("[error]Datum kraja ne može biti veći od datuma početka. Pokušajte ponovo.[/]")
        else:
            break

    date_range = []
    if int(date_from) == int(date_to):
        date_range.append(int(date_from))
    else:
        date_range = list(range(int(date_from), int(date_to) + 1))

    if choice.upper() == "BO":
        for day in date_range:
            day = int(day) + 5
            # Select the cell to check
            fieldH = "H" + str(day)
            cell = worksheet[fieldH]

            # Check if the cell is empty or not
            if cell.value is None:
                colB = "B" + str(day)
                colC = "C" + str(day)
                colD = "D" + str(day)
                colE = "E" + str(day)
                colF = "F" + str(day)
                colG = "G" + str(day)
                colH = "H" + str(day)

                # Write to the worksheet
                worksheet[colB] = ''
                worksheet[colC] = ''
                worksheet[colD] = ''
                worksheet[colE] = '{:02d}:00'.format(rv_pocetak).zfill(5)
                worksheet[colF] = '{:02d}:00'.format(rv_kraj).zfill(5)
                worksheet[colG] = '{:02d}:00'.format(rv_sveukupno).zfill(5)
                worksheet[colH] = choice.upper()

                # Save the workbook to a file
                workbook.save(document)
            else:
                console.print(str(day) + ". je " + str(cell.value), style="warning")

    elif choice.upper() == "GO":
        for day in date_range:
            day = int(day) + 5
            # Select the cell to check
            fieldH = "H" + str(day)
            cell = worksheet[fieldH]

            # Check if the cell is empty or not
            if cell.value is None:
                colB = "B" + str(day)
                colC = "C" + str(day)
                colD = "D" + str(day)
                colE = "E" + str(day)
                colF = "F" + str(day)
                colG = "G" + str(day)
                colH = "H" + str(day)

                # Write to the worksheet
                worksheet[colB] = ''
                worksheet[colC] = ''
                worksheet[colD] = ''
                worksheet[colE] = '{:02d}:00'.format(rv_pocetak).zfill(5)
                worksheet[colF] = '{:02d}:00'.format(rv_kraj).zfill(5)
                worksheet[colG] = '{:02d}:00'.format(rv_sveukupno).zfill(5)
                worksheet[colH] = choice.upper()

                # Save the workbook to a file
                workbook.save(document)
            else:
                console.print(str(day) + ". je " + str(cell.value), style="warning")

    elif choice.upper() == "SP":
        for day in date_range:
            day = int(day) + 5
            # Select the cell to check
            fieldH = "H" + str(day)
            cell = worksheet[fieldH]

            # Check if the cell is empty or not
            if cell.value is None:
                colE = "E" + str(day)
                colF = "F" + str(day)
                colG = "G" + str(day)
                colH = "H" + str(day)

                # Write to the worksheet
                worksheet[colE] = '{:02d}:00'.format(rv_pocetak).zfill(5)
                worksheet[colF] = '{:02d}:00'.format(rv_kraj).zfill(5)
                worksheet[colG] = '{:02d}:00'.format(rv_sveukupno).zfill(5)
                worksheet[colH] = choice.upper()

                # Save the workbook to a file
                workbook.save(document)
            else:
                console.print(str(day) + ". je " + str(cell.value), style="warning")
    else:
        ask_for_custom(document)


def ask_for_custom(document):
    while True:
        # Clearing console
        os.system('cls' if os.name == 'nt' else 'clear')

        table = Table(title="Odaberi...")

        table.add_column("OZNAKA", style="cyan")
        table.add_column("Opis oznake", style="cyan")

        table.add_row("BO", "bolovanje")
        table.add_row("BR", "rad blagdanom")
        table.add_row("DO", "dnevni odmor")
        table.add_row("GO", "godišnji odmor")
        table.add_row("LO", "lockout")
        table.add_row("ND", "neplaćeni dopust")
        table.add_row("NR", "noćni rad")
        table.add_row("ODS", "odsutnost")
        table.add_row("PD", "plaćeni dopust")
        table.add_row("PR", "prekovremeni rad")
        table.add_row("PRE", "preraspodjela")
        table.add_row("PREK", "prekid - poslodavac")
        table.add_row("PRI", "pripravnost")
        table.add_row("ROD", "po(rodiljni) dopust")
        table.add_row("SP", "službeni put")
        table.add_row("TER", "terenski rad")
        table.add_row("TO", "tjedni odmor")
        table.add_row("ZAS", "zastoj")
        table.add_row("EXIT", "ako želiš izaći.")

        console.print(table)
        choice = input("Unesite inicijale (npr. GO za Godišnji Odmor): ")

        # Get the selected month based on the user's choice
        if choice.upper() == "BO" or choice.upper() == "GO" or choice.upper() == "SP":
            insert_custom(document, choice)
        elif choice.upper() == "EXIT":
            break
        else:
            console.print(choice + " još nije aktivan. Odaberi drugi.", style="error")


def we_are_done_here(file_name, selected_month_converted, year):
    # Clearing console
    os.system('cls' if os.name == 'nt' else 'clear')

    while os.path.exists(file_name):
        console.print("[cyan]Evidencija rada za[/] [warning]" + str(selected_month_converted) + " " + str(year) + "[/] [cyan]je izrađena.[/]")
        console.print(" ", style="black")
        console.print("Dokument je moguće pronaći ovdje:")
        console.print("[success]"+file_name+"[/]")
        console.print(" ", style="black")
        input("Press Enter to continue and open Document...")
        os.startfile(file_name)
        exit()


def create_new_document(name, year, selected_month, selected_month_converted):
    # Clearing console
    os.system('cls' if os.name == 'nt' else 'clear')

    if zaposlenik is None or zaposlenik == "":
        worker = input("Unesite ime i prezime radnika: ")
    else:
        worker = str(zaposlenik)

    # Create or load an existing workbook
    # workbook = openpyxl.Workbook()
    # If you want to open an existing workbook, replace the above line with the one below
    workbook = openpyxl.load_workbook(name)

    # Access the active worksheet
    worksheet = workbook.active

    # Write to the worksheet
    worksheet['C1'] = worker.upper()
    worksheet['B2'] = year
    worksheet['F2'] = selected_month_converted
    worksheet['D39'] = todays_date

    # Save the workbook to a file
    file_name = str(year) + "_" + str(selected_month_converted) + "_" + str(worker.upper().replace(" ", "_")) + "_" + str(datetime.now().strftime('%H%M%S%f')) + ".xlsx"
    file_name = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop', file_name)
    workbook.save(file_name)

    # Fill new document with data
    console.print("[success]Unosim radne dane...[/]")
    working_day_hours = content_filler(file_name)

    # Insert Weekends
    console.print("[success]Unosim vikende...[/]")
    clean_weekends(file_name, get_sundays(year, selected_month))

    # Insert Holidays
    console.print("[success]Unosim praznike i blagdane dane...[/]")
    clean_holidays(file_name, get_holidays(year, selected_month))

    # Ask for custom inputs
    ask_for_custom(file_name)

    # Count UKUPNO Hours
    console.print("[success]Izračunavam sveukupno...[/]")
    count_ukupno(file_name, working_day_hours)

    # Creation msg
    we_are_done_here(file_name, selected_month_converted, year)


def ask_for_month():
    while True:
        # Display a menu for the user to select the month
        console.print("[warning]Odaberite period izrade:[/]")
        console.print("[general]➡️[/] [success](1)[/] Prethodni mjesec [cyan]("+month_converter(current_month)+")[/]")
        console.print("[general]➡️[/] [success](2)[/] Aktualni mjesec [cyan]("+month_converter(previous_month)+")[/]")
        console.print("[general]➡️[/] [success](3)[/] Želim sam/a odabrati")
        console.print(" ", style="black")
        choice = int(input("Odaberi (1, 2 ili 3): "))

        # Get the selected month based on the user's choice
        if choice == 1:
            create_new_document(DOCUMENT_TEMPLATE, last_month_year, previous_month, month_converter(previous_month))
            break
        elif choice == 2:
            create_new_document(DOCUMENT_TEMPLATE, last_month_year, current_month, month_converter(current_month))
            break
        elif choice == 3:
            # Clearing console
            os.system('cls' if os.name == 'nt' else 'clear')

            custom_year = int(input("Unesi godinu (npr. 2023): "))
            while len(str(custom_year)) != 4:
                custom_year = int(input("Neispravan unos, pokušajte ponovo: "))
            print("Odabrana godina: ", custom_year)

            custom_month = int(input("Unesi mjesec (npr. 3): "))
            while not custom_month <= 12:
                custom_month = int(input("Neispravan unos, pokušajte ponovo: "))
            print("Odabrani mjesec: ", month_converter(custom_month))

            create_new_document(DOCUMENT_TEMPLATE, custom_year, custom_month, month_converter(custom_month))
            break
        else:
            print("Krivi odabir! Pokušaj ponovo.")


def welcome_msg():
    # Clearing console
    os.system('cls' if os.name == 'nt' else 'clear')

    console.print(" ", style="black")
    console.print("🕒 IZRADA EVIDENCIJE RADNOG VREMENA 🕒", style="bold cyan")
    console.print("[general]📅 Datum:[/] [accent]" + todays_date + "[/]")
    console.print(" ", style="black")


if __name__ == '__main__':
    welcome_msg()
    ask_for_month()

